#!/usr/bin/env python3
# pylint: disable=too-many-lines
"""Compare cell lists from two process libraries and write an Excel report.

Usage:
    python3 compare_cells.py --function-keys-file preview/function_key_zh.txt
    python3 compare_cells.py --function-keys OAI2211 OAI22 AN2 AN3

One KEY<TAB>中文 file is enough: column 1 is match order, column 2 is the
Chinese label shown in Excel column A (edit the file to change labels).

Column A uses the first table key that is a qualified prefix of the cell
name (order matters: put longer forms first). Matched cells are not
reconsidered by later keys.

Only identical full cell names share a data row; NP-only / C1-only cells
each get their own row under the shared function key.
"""

from __future__ import annotations

import argparse
import os
import re
import shlex
import sys
from collections import defaultdict
from typing import Dict, List, Optional, Pattern, Sequence, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# #### Constants
NP_FILE = "/mnt/c/Users/t00961128/Downloads/NP1PP.list"
C1_FILE = "/mnt/c/Users/t00961128/Downloads/C1Y.list"
OUTPUT = "/mnt/c/Users/t00961128/Downloads/NP1PP_vs_C1Y.xlsx"

SHEET_TITLE = "Cell Comparison"
COL_WIDTHS = {"A": 28, "B": 55, "C": 55}
FONT_SIZE = 16
MATCH_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
NORMAL_FONT = Font(size=FONT_SIZE)
BOLD_FONT = Font(size=FONT_SIZE, bold=True)
KEY_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER_SIDE = Side(style="thin", color="000000")
TOP_BOTTOM_BORDER = Border(top=BORDER_SIDE, bottom=BORDER_SIDE)
TOP_BORDER = Border(top=BORDER_SIDE)
BOTTOM_BORDER = Border(bottom=BORDER_SIDE)
REPLACE_SEP_RE = re.compile(r"\s*:\s*")

# Mid-name process/optimization markers: cut from here to end (pre-COT).
FUNCTION_CUT_TOKENS = (
    "EEQMBD",
    "EEQMBC",
    "EEQ",
    "OPT",
    "SKFMDB",
    "SKRMDB",
    "SKND",
    "SKPD",
    "SKRD",
    "SKF",
    "SKR",
    "SKN",
    "SKP",
    "TWA",
)

# Layout / track suffixes peeled only when preceded by a digit.
_TRAILING_TAG_NAMES = (
    "CNWBAL",
    "CNW",
    "CWBAL",
    "CWRB",
    "BALRB",
    "BAL",
    "RBD",
    "AROAF",
    "3FINC",
    "3FIN",
    "SHXP",
    "DHXP",
    "XP",
    "VG",
    "SH",
    "DH",
    "SF",
    "BF",
    "FF",
    "AF",
    "AR2",
    "AR1",
    "ARP",
    "ARN",
    "ARQ",
    "ARO",
    "AAR",
    "AR",
    "APA",
    "APB",
    "APC",
    "APD",
    "APM",
    "APP",
    "APN",
    "APQ",
    "ANS",
    "AKP",
    "AKN",
    "ATA",
    "ATB",
    "ATC",
    "ATS",
    "ATF",
    "ALP",
    "AHB",
    "AHA",
    "AHQ",
    "AHC",
    "ASB",
    "ONE",
    "ME",
    "ICC",
    "LN",
    "TG",
    "P5",
    "SC4LP",
    "SC5LP",
    "COMB1",
)
TRAILING_TAG_RE = re.compile(
    r"(?<=\d)(?:"
    + "|".join(
        re.escape(tag)
        for tag in sorted(_TRAILING_TAG_NAMES, key=len, reverse=True)
    )
    + r")$"
)
DRIVE_RE = re.compile(r"[DX]\d+(?:P\d+)?$")
# Whole-name roots that still end with D\d+/X\d+ and must not peel.
PROTECTED_WHOLE_RE = re.compile(
    r"(?:"
    r"(?:CK|G)?(?:ND|NR|IND|INR|IIND|IINR|GND|GNR|GNAND|GNOR)\d+|"
    r"(?:G)?AND\d+|"
    r"(?:CK|G)?MUX\d+|"
    r"MX\d+"
    r")$"
)
# Real compound *ND2+/*NR2+ (optional extra N: MUX2NND2); not ND1 drive.
COMPOUND_ND_NR_RE = re.compile(
    r"(?:AOI|OAI|XOR|XNR|MUX|AN|OR|AO|OA|ND|NR|INR|IND)\d+"
    r"N?(?:ND|NR)(?:[2-9]\d*)$"
)
# Non-digit-prefixed variants peeled after drive strength.
VARIANT_SUFFIX_RE = re.compile(
    r"(?:CCB|CCM|CCA|SNK|SRC|CW|CWBAL|CWRB|BALRB|BAL|DBA4|NOBCM|"
    r"TGAR|XNRAR|ARSP|VPPVBB|IW|V2|COM|XP|XN|FR)$"
)
# Device/ratio variants after topology digits, e.g. AOI21B1 / AOI21N2 / ND2N1.
# Require a preceding digit so AN2 / GAN2 are not peeled.
BN_VARIANT_RE = re.compile(r"(?<=\d)[BN]\d$")
# Level-shifter domain side after LH/HL direction (not a distinct logic function).
LVL_SIDE_RE = re.compile(r"(?<=(?:LH|HL))(?:CH|CL)$")
# "HD" layout marker; keep short roots like BHD (stem shorter than 4).
HD_SUFFIX_MIN_STEM = 4
# Physical layout families: collapse sized variants to one key each.
LAYOUT_FAMILY_PREFIXES = (
    "BOUNDARY",
    "HDDICWY",
    "HDDID",
)
COT_AND_AFTER_RE = re.compile(r"COT.*$")
# Flip-flop family stems (longest first): keep through stem, drop Q/RPQ/SNQ/...
FF_STEMS = (
    "Y3SDFF",
    "Y2SDFF",
    "Y3SDF",
    "Y2SDF",
    "YSDF",
    "SEDF",
    "RSDF",
    "GSDF",
    "SD2FF",
    "SDFF",
    "SDF",
    "GDF",
    "EDF",
    "DF",
)

CellRow = Tuple[str, Optional[str], Optional[str]]
GroupItem = Tuple[str, bool, bool]
ReplaceRule = Tuple[Pattern[str], str]


# #### CLI
def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description=(
            "Compare NP1PP and C1Y cell lists by ordered function-key "
            "prefix match (first hit wins)."
        ),
    )
    parser.add_argument(
        "--function-keys",
        nargs="+",
        default=[],
        metavar="KEY",
        dest="function_keys",
        help=(
            "Ordered function keys to keep as column A. First prefix match "
            "wins; later keys do not override. A digit right after the key "
            "blocks the match (OAI22 will not take OAI2211). Still put "
            "longer letter-tailed forms first, e.g. "
            "--function-keys OAI22OAI21 OAI22 AIOI21 AN2 AN3 AN4."
        ),
    )
    parser.add_argument(
        "--function-keys-file",
        default="",
        metavar="PATH",
        dest="function_keys_file",
        help=(
            "Ordered key table file. Prefer KEY<TAB>中文 (one pair per "
            "line), which alone supplies both matching order and Chinese "
            "labels (e.g. preview/function_key_zh.txt). Plain one-token "
            "keys per line are also accepted. Appended after "
            "--function-keys; first-match-wins."
        ),
    )
    parser.add_argument(
        "--function-key-zh-file",
        default="",
        metavar="PATH",
        dest="function_key_zh_file",
        help=(
            "Same KEY<TAB>中文 table as --function-keys-file. Use this "
            "alone when the table is your only config; or use it to "
            "override Chinese when keys come from elsewhere. Missing "
            "Chinese falls back to built-in describe_function_key()."
        ),
    )
    parser.add_argument(
        "--dump-suggested-keys",
        default="",
        metavar="PATH",
        dest="dump_suggested_keys",
        help=(
            "Write a longest-first KEY<TAB>中文 table from the input cell "
            "names to PATH (single file; edit the Chinese column later), "
            "then exit."
        ),
    )
    parser.add_argument(
        "--format-filter",
        default="",
        metavar="TOKENS",
        dest="format_filter",
        help=(
            "Optional quote-delimited regex tokens to delete before "
            "prefix matching, applied left-to-right, "
            "e.g. '\"FOO\" \"BAR\"'."
        ),
    )
    parser.add_argument(
        "--format-replace",
        default="",
        metavar="MAPPINGS",
        dest="format_replace",
        help=(
            "Quote-delimited regex replacements applied after deletes, "
            'e.g. \'"FOO : BAR" "X(\\d{1,2}) : D\\\\1"\'. '
            "Use ASCII ':' between pattern and replacement."
        ),
    )
    parser.add_argument(
        "--np-file",
        default=NP_FILE,
        help=f"Path to NP1PP list file (default: {NP_FILE})",
    )
    parser.add_argument(
        "--c1-file",
        default=C1_FILE,
        help=f"Path to C1Y list file (default: {C1_FILE})",
    )
    parser.add_argument(
        "--output",
        default=OUTPUT,
        help=f"Output Excel path (default: {OUTPUT})",
    )
    return parser.parse_args(argv)


def split_quoted_tokens(raw: str, option_name: str) -> List[str]:
    """Split a quote-delimited option value into tokens via shlex."""
    if not raw or not raw.strip():
        return []
    try:
        tokens = shlex.split(raw, posix=True)
    except ValueError as exc:
        print(f"Invalid {option_name} quoting: {exc}", file=sys.stderr)
        sys.exit(2)
    return [token for token in tokens if token]


def split_format_filter(raw: str) -> List[str]:
    """Split --format-filter into quote-delimited regex tokens."""
    return split_quoted_tokens(raw, "--format-filter")


def parse_replace_mapping(token: str) -> Tuple[str, str]:
    """Split one replace token into (pattern, replacement) on ASCII ':'."""
    match = REPLACE_SEP_RE.search(token)
    if match is None:
        print(
            "Invalid --format-replace token "
            f"(missing ':'): {token!r}",
            file=sys.stderr,
        )
        sys.exit(2)
    pattern = token[: match.start()].strip()
    replacement = token[match.end() :].strip()
    if not pattern:
        print(
            f"Invalid --format-replace token (empty pattern): {token!r}",
            file=sys.stderr,
        )
        sys.exit(2)
    return pattern, replacement


def parse_format_replace(raw: str) -> List[Tuple[str, str]]:
    """Parse --format-replace into ordered (pattern, replacement) pairs."""
    mappings: List[Tuple[str, str]] = []
    for token in split_quoted_tokens(raw, "--format-replace"):
        mappings.append(parse_replace_mapping(token))
    return mappings


def load_function_key_table(
    filepath: str,
) -> Tuple[List[str], Dict[str, str]]:
    """Load ordered keys and optional Chinese from one table file.

    Accepted lines (comments / blanks skipped):
    - ``KEY<TAB>中文`` — preferred; one file for match order + labels
    - ``KEY`` — key only; Chinese falls back to built-in rules later
    """
    keys: List[str] = []
    mapping: Dict[str, str] = {}
    with open(filepath, encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            if "\t" in stripped:
                key, chinese = stripped.split("\t", 1)
                key = key.strip()
                chinese = chinese.strip()
                if not key or any(char.isspace() for char in key):
                    print(
                        f"Invalid key in {filepath}:{line_number}: "
                        f"{stripped!r} (KEY must be a single token)",
                        file=sys.stderr,
                    )
                    sys.exit(2)
                keys.append(key)
                if chinese:
                    mapping[key] = chinese
                continue
            if any(char.isspace() for char in stripped):
                print(
                    f"Invalid key in {filepath}:{line_number}: {stripped!r} "
                    "(use KEY or KEY<TAB>中文 per line)",
                    file=sys.stderr,
                )
                sys.exit(2)
            keys.append(stripped)
    return keys, mapping


def load_function_keys_file(filepath: str) -> List[str]:
    """Load ordered function keys (wrapper; prefer load_function_key_table)."""
    keys, _mapping = load_function_key_table(filepath)
    return keys


def load_function_key_zh_file(filepath: str) -> Dict[str, str]:
    """Load KEY -> Chinese map (wrapper; prefer load_function_key_table)."""
    _keys, mapping = load_function_key_table(filepath)
    return mapping


def resolve_function_key_config(
    cli_keys: Sequence[str],
    keys_file: str,
    zh_file: str,
) -> Tuple[List[str], Dict[str, str], str]:
    """Resolve ordered keys + Chinese map from CLI and optional files.

    One ``KEY<TAB>中文`` file is enough: pass it as ``--function-keys-file``
    or ``--function-key-zh-file``. Returns (keys, zh_map, table_path_note).
    """
    keys = [key for key in cli_keys if key]
    zh_map: Dict[str, str] = {}
    table_note = ""

    if keys_file:
        file_keys, file_zh = load_function_key_table(keys_file)
        keys.extend(file_keys)
        zh_map.update(file_zh)
        table_note = keys_file

    if zh_file:
        file_keys, file_zh = load_function_key_table(zh_file)
        if not keys:
            keys.extend(file_keys)
        zh_map.update(file_zh)
        table_note = zh_file
    elif keys_file and not zh_map:
        sibling = zh_path_for_keys_file(keys_file)
        if os.path.isfile(sibling):
            _unused_keys, file_zh = load_function_key_table(sibling)
            zh_map.update(file_zh)
            table_note = sibling

    if not keys:
        print(
            "Error: provide --function-keys, --function-keys-file, and/or "
            "--function-key-zh-file with at least one key "
            "(KEY<TAB>中文 in one file is enough).",
            file=sys.stderr,
        )
        sys.exit(2)
    return keys, zh_map, table_note


def zh_path_for_keys_file(keys_path: str) -> str:
    """Derive function_key_zh.txt path beside a function_keys.txt path."""
    directory, filename = os.path.split(keys_path)
    stem, ext = os.path.splitext(filename)
    if stem.endswith("_keys"):
        zh_name = stem[: -len("_keys")] + "_key_zh" + (ext or ".txt")
    elif stem == "function_keys":
        zh_name = "function_key_zh" + (ext or ".txt")
    else:
        zh_name = stem + "_zh" + (ext or ".txt")
    return os.path.join(directory, zh_name) if directory else zh_name


# #### I/O
def read_list(filepath: str) -> List[str]:
    """Read a cell list file, skipping blanks and rg: header lines."""
    items: List[str] = []
    with open(filepath, encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if not stripped or stripped.startswith("rg:"):
                continue
            items.append(stripped)
    return items


# #### Name helpers
def compile_regex_filters(patterns: Sequence[str]) -> List[Pattern[str]]:
    """Compile regex filter patterns, exiting on invalid syntax."""
    compiled: List[Pattern[str]] = []
    for pattern in patterns:
        try:
            compiled.append(re.compile(pattern))
        except re.error as exc:
            print(
                f"Invalid regex in --format-filter: {pattern!r}: {exc}",
                file=sys.stderr,
            )
            sys.exit(2)
    return compiled


def compile_regex_replaces(
    mappings: Sequence[Tuple[str, str]],
) -> List[ReplaceRule]:
    """Compile regex replacement rules, exiting on invalid syntax."""
    compiled: List[ReplaceRule] = []
    for pattern, replacement in mappings:
        try:
            compiled.append((re.compile(pattern), replacement))
        except re.error as exc:
            print(
                f"Invalid regex in --format-replace: {pattern!r}: {exc}",
                file=sys.stderr,
            )
            sys.exit(2)
    return compiled


def apply_regex_filters(
    value: str,
    regex_filters: Sequence[Pattern[str]],
) -> str:
    """Apply delete regexes left-to-right; each runs until stable."""
    for pattern in regex_filters:
        while True:
            updated = pattern.sub("", value)
            if updated == value:
                break
            value = updated
    return value


def apply_regex_replaces(
    value: str,
    regex_replaces: Sequence[ReplaceRule],
) -> str:
    """Apply replace regexes left-to-right once each via re.sub."""
    for pattern, replacement in regex_replaces:
        value = pattern.sub(replacement, value)
    return value


def _cut_process_tail(name: str) -> str:
    """Truncate at the earliest mid-name process/optimization marker."""
    cut_at = len(name)
    for token in FUNCTION_CUT_TOKENS:
        index = name.find(token)
        if index != -1 and index < cut_at:
            cut_at = index
    return name[:cut_at]


def _peel_trailing_tags(name: str) -> str:
    """Peel layout/track suffixes that appear immediately after a digit."""
    while True:
        match = TRAILING_TAG_RE.search(name)
        if match is None:
            break
        name = name[: match.start()]
    return name


def _peel_variant_suffixes(name: str) -> str:
    """Peel known non-drive variant suffixes (CCA/CCB/SNK/HD/B1/N2/...)."""
    while True:
        match = VARIANT_SUFFIX_RE.search(name)
        if match is not None:
            name = name[: match.start()]
            continue
        if name.endswith("HD") and len(name) - 2 >= HD_SUFFIX_MIN_STEM:
            name = name[:-2]
            continue
        match = BN_VARIANT_RE.search(name)
        if match is not None:
            name = name[: match.start()]
            continue
        match = LVL_SIDE_RE.search(name)
        if match is not None:
            name = name[: match.start()]
            continue
        break
    return name


def _should_stop_drive_peel(name: str) -> bool:
    """True when trailing D/X digits are part of the function, not drive."""
    if PROTECTED_WHOLE_RE.fullmatch(name):
        return True
    if COMPOUND_ND_NR_RE.search(name):
        return True
    return False


def _normalize_layout_family(name: str) -> str:
    """Collapse sized layout cells to their family prefix key."""
    for prefix in LAYOUT_FAMILY_PREFIXES:
        if name.startswith(prefix):
            return prefix
    return name


def _normalize_function_aliases(name: str) -> str:
    """Merge remaining non-functional aliases into logic roots."""
    # ISOCH/ISOCL domain-side C → ISOH/ISOL.
    if name.startswith("ISOC") and len(name) >= 5 and name[4] in "HL":
        name = "ISO" + name[4:]
    # LVUFR* mid-token FR variant → LVU*.
    if name.startswith("LVUFR"):
        name = "LVU" + name[5:]
    # Level-shifter buffered variant → same direction root.
    if re.match(r"^(?:CK)?LVL.*BUFF$", name):
        name = name[:-4]
    if name == "BUFF":
        name = "BUF"
    return name


def _collapse_ff_tail(name: str) -> str:
    """Omit Q/RPQ/SNQ/... after DF/SDF/EDF/SDFF/... flip-flop stems."""
    best_index: Optional[int] = None
    best_end: Optional[int] = None
    for stem in FF_STEMS:
        index = name.find(stem)
        if index == -1:
            continue
        end = index + len(stem)
        if (
            best_index is None
            or index < best_index
            or (index == best_index and end > best_end)
        ):
            best_index = index
            best_end = end
    if best_end is None:
        return name
    return name[:best_end]


def extract_function_root(name: str) -> str:
    """Extract a suggested logic-function root from a full cell name.

    Used to build default key lists; runtime grouping uses --function-keys
    prefix matching instead.
    """
    value = COT_AND_AFTER_RE.sub("", name)
    value = _cut_process_tail(value)
    while True:
        value = _peel_trailing_tags(value)
        value = _peel_variant_suffixes(value)
        if _should_stop_drive_peel(value):
            break
        match = DRIVE_RE.search(value)
        if match is None:
            break
        value = value[: match.start()]
    value = _peel_trailing_tags(value)
    value = _peel_variant_suffixes(value)
    value = _collapse_ff_tail(value)
    value = _normalize_layout_family(value)
    return _normalize_function_aliases(value)


def match_function_key(
    name: str,
    function_keys: Sequence[str],
) -> Optional[str]:
    """Return the first matching function key for name, else None.

    Prefer the canonical extract_function_root when it appears in the key
    list (so ISOCH… maps to ISOH after aliasing). Otherwise fall back to
    ordered prefix match: earlier keys win permanently; a digit right after
    the key blocks the match except for layout family keys.
    """
    root = extract_function_root(name)
    if root in set(function_keys):
        return root

    layout_families = set(LAYOUT_FAMILY_PREFIXES)
    for key in function_keys:
        if not name.startswith(key):
            continue
        rest = name[len(key) :]
        if rest and rest[0].isdigit() and key not in layout_families:
            continue
        return key
    return None


def describe_function_key(  # pylint: disable=too-many-return-statements,too-many-branches,too-many-statements
    key: str,
) -> str:
    """Return a short Chinese label for a function key (may be empty)."""
    if key.startswith(("BOUNDARY", "HDDICWY", "HDDID")):
        return "边界/版图单元"
    if key.startswith("ANTENNA"):
        return "天线二极管"
    if key.startswith("TAPCELL"):
        return "阱敲击单元"
    if key.startswith(("FILL", "GFILL")):
        return "填充单元"
    if key.startswith(("DCAP", "GDCAP")):
        return "去耦电容"
    if key.startswith(("TIE", "GTIE", "APTIE")):
        return "电平钳位"
    if key == "BHD":
        return "总线保持器"
    if key.startswith("DEL"):
        return "延时单元"
    if "PULL" in key:
        return "上拉/下拉"
    if key.startswith("ISO"):
        return "隔离单元"
    if key.startswith(("LVL", "LVU", "CKLVL", "CKLV")):
        return "电平转换器"
    if key.startswith("SYN"):
        return "同步器"
    if key == "BUFT":
        return "三态缓冲器"
    if key == "INVPAD":
        return "焊盘反相器"
    if key.startswith(("BUFF", "BUF", "GBUFF", "APBUF")):
        return "缓冲器"
    if re.match(r"^INV[A-Z]", key) and ("AOI" in key or "OAI" in key):
        return "复合逻辑门"
    if key.startswith(("INV", "GINV", "APINV")):
        return "反相器"
    if key.startswith(("CK", "DCCK", "GCK")):
        if "MUX" in key:
            return "时钟多路选择器"
        if re.search(r"ND\d", key) or "NAND" in key:
            return "时钟与非门"
        if re.search(r"NR\d", key) or "NOR" in key:
            return "时钟或非门"
        if re.search(r"AN\d", key):
            return "时钟与门"
        if "XOR" in key:
            return "时钟异或门"
        if re.search(r"OR\d", key):
            return "时钟或门"
        if key in ("CKB", "DCCKB", "CKN") or key.endswith("CKB"):
            return "时钟缓冲/反相"
        if any(token in key for token in ("LNQ", "LHQ", "LNCNQ", "LHCNQ")):
            return "时钟锁存器"
        if key.startswith("GCK"):
            return "门控时钟"
        return "时钟单元"
    if key.startswith("MB") or key.startswith("MCE"):
        if any(stem in key for stem in ("SDF", "DF", "SDFF", "SD2FF")):
            return "多比特触发器"
        if "LH" in key or "LN" in key or "CNQ" in key:
            return "多比特锁存器"
        return "多比特触发器"
    if key in {"EDF"} or key.startswith("EDF"):
        return "使能触发器"
    if key in {"DF", "GDF"} or key.startswith("GDF"):
        return "D触发器"
    if key in {
        "SDF",
        "SDFF",
        "SD2FF",
        "SEDF",
        "RSDF",
        "GSDF",
        "YSDF",
        "Y2SDF",
        "Y3SDF",
        "Y2SDFF",
        "Y3SDFF",
    } or any(
        key.startswith(prefix)
        for prefix in (
            "SDF",
            "SDFF",
            "SD2FF",
            "SEDF",
            "RSDF",
            "GSDF",
            "YSDF",
            "Y2SDF",
            "Y3SDF",
        )
    ):
        return "扫描触发器"
    if key.startswith(("DF", "GDF")):
        return "D触发器"
    if any(
        key.startswith(prefix)
        for prefix in (
            "LHQ",
            "LNQ",
            "LHCNQ",
            "LNCNQ",
            "LHSNQ",
            "LNSNQ",
            "LHCSNQ",
            "LNCSNQ",
            "GLHQ",
            "GLNQ",
            "GLAHQ",
            "GLHCNQ",
            "GLNCNQ",
        )
    ):
        return "锁存器"
    if key.startswith("FA"):
        return "全加器"
    if key.startswith(("HA", "HAC")):
        return "半加器"
    if key.startswith("CMPE"):
        return "压缩器"
    if key.startswith(("FC", "FI", "FCON")):
        return "进位逻辑"
    if key.startswith("BENC"):
        return "编码器"
    if "MUX" in key or key.startswith(("MXI", "MX", "GMUX")):
        if any(token in key for token in ("ND", "NR", "AOI", "NND", "NNR")):
            return "复合多路选择器"
        if (
            key.startswith("MXI")
            or re.search(r"MUX\d+I", key)
            or re.search(r"MUX\d+N$", key)
        ):
            return "反相多路选择器"
        return "多路选择器"
    if key.startswith(
        (
            "MAOI",
            "MOAI",
            "AOAI",
            "OAOI",
            "AIOI",
            "IAO",
            "IAOI",
            "IOA",
            "IOAI",
            "OIAI",
            "WAO",
            "W2AO",
        )
    ):
        return "复合逻辑门"
    if re.match(r"^(XOR|XNR|AOI|OAI|ND|NR|AN|OR)\d+", key):
        rest = re.sub(r"^(XOR|XNR|AOI|OAI|ND|NR|AN|OR)\d+", "", key)
        if rest and re.match(
            r"^(AOI|OAI|XOR|XNR|ND|NR|INV|AN|OR|AO|OA|IOA)",
            rest,
        ):
            return "复合逻辑门"
    if key.startswith(("AOI", "GAOI")):
        return "与或非门"
    if key.startswith(("OAI", "GOAI")):
        return "或与非门"
    if key.startswith(("AO", "GAO")) and not key.startswith("AOI"):
        return "与或门"
    if key.startswith(("OA", "GOA")) and not key.startswith("OAI"):
        return "或与门"
    if key.startswith(("XOR", "GXOR")):
        return "异或门"
    if key.startswith(("XNR", "GXNR")):
        return "同或门"
    if re.match(r"^(G)?(ND|NAND|IND|IIND|GNAND|GND|GIND)\d", key) or re.match(
        r"^ND\d",
        key,
    ):
        return "与非门"
    if re.match(r"^(G)?(NR|NOR|INR|IINR|GNOR|GNR)\d", key) or re.match(
        r"^NR\d",
        key,
    ):
        return "或非门"
    if (
        re.match(r"^(G)?AN\d", key)
        or re.match(r"^GAND", key)
        or re.match(r"^GAN\d", key)
    ):
        return "与门"
    if re.match(r"^(G)?OR\d", key) or re.match(r"^GOR\d", key):
        return "或门"
    if key.startswith("AP"):
        return "辅助单元"
    return ""


def format_column_a_label(
    function_key: str,
    zh_map: Dict[str, str] | None = None,
) -> str:
    """Build merged-cell text: function key plus Chinese brief."""
    chinese = ""
    if zh_map and function_key in zh_map:
        chinese = zh_map[function_key]
    if not chinese:
        chinese = describe_function_key(function_key)
    if chinese:
        return f"{function_key}\n{chinese}"
    return function_key


def make_display_key(
    name: str,
    function_keys: Sequence[str],
    regex_filters: Sequence[Pattern[str]],
    regex_replaces: Sequence[ReplaceRule] | None = None,
) -> Tuple[str, bool]:
    """Build column-A key via ordered prefix match.

    Returns (key, matched). Unmatched cells keep the post-filter name so
    gaps stay visible in the spreadsheet.
    """
    value = apply_regex_filters(name, regex_filters)
    if regex_replaces:
        value = apply_regex_replaces(value, regex_replaces)
    matched = match_function_key(value, function_keys)
    if matched is not None:
        return matched, True
    return value, False


# #### Grouping
def group_cells(  # pylint: disable=too-many-arguments,too-many-positional-arguments
    all_cells: Sequence[str],
    np_set: Set[str],
    c1_set: Set[str],
    function_keys: Sequence[str],
    regex_filters: Sequence[Pattern[str]],
    regex_replaces: Sequence[ReplaceRule] | None = None,
) -> Tuple[Dict[str, List[GroupItem]], List[str]]:
    """Group cells by function key; return groups and unmatched names."""
    groups: Dict[str, List[GroupItem]] = defaultdict(list)
    unmatched: List[str] = []
    for cell in all_cells:
        key, matched = make_display_key(
            cell,
            function_keys,
            regex_filters,
            regex_replaces,
        )
        if not matched:
            unmatched.append(cell)
        groups[key].append((cell, cell in np_set, cell in c1_set))
    return groups, unmatched


def dump_suggested_keys(
    cells: Sequence[str],
    output_path: str,
) -> Tuple[List[str], str]:
    """Write one longest-first KEY\\t中文 table (match order + labels)."""
    roots = {extract_function_root(cell) for cell in cells}
    keys = sorted(root for root in roots if root)
    keys.sort(key=lambda item: (-len(item), item))
    with open(output_path, "w", encoding="utf-8") as handle:
        handle.write("# key\tchinese\n")
        handle.write(
            "# One file for both: column1 = match order, column2 = A-column "
            "Chinese (edit freely).\n"
        )
        handle.write(f"# cells={len(cells)} keys={len(keys)}\n")
        for key in keys:
            handle.write(f"{key}\t{describe_function_key(key)}\n")
    return keys, output_path


def _partition_group(
    items: Sequence[GroupItem],
) -> Tuple[List[str], List[str], List[str]]:
    """Split a group into exact matches, NP-only, and C1-only lists."""
    exact_matches: List[str] = []
    np_only: List[str] = []
    c1_only: List[str] = []
    for cell, in_np, in_c1 in items:
        if in_np and in_c1:
            exact_matches.append(cell)
        elif in_np:
            np_only.append(cell)
        elif in_c1:
            c1_only.append(cell)
    exact_matches.sort()
    np_only.sort()
    c1_only.sort()
    return exact_matches, np_only, c1_only


def build_rows(
    groups: Dict[str, List[GroupItem]],
    zh_map: Dict[str, str] | None = None,
) -> List[CellRow]:
    """Build Excel data rows from grouped cells.

    Only identical full cell names share a row (both NP and C1 filled).
    NP-only and C1-only cells each get their own row; they are never
    zip-paired across libraries. Column A shows function key plus a short
    Chinese description.
    """
    rows: List[CellRow] = []
    for key in sorted(groups.keys()):
        display = format_column_a_label(key, zh_map)
        exact_matches, np_only, c1_only = _partition_group(groups[key])
        for cell in exact_matches:
            rows.append((display, cell, cell))
        for cell in np_only:
            rows.append((display, cell, None))
        for cell in c1_only:
            rows.append((display, None, cell))
    return rows


# #### Excel writing
def _iter_display_key_groups(rows: Sequence[CellRow]) -> List[Tuple[int, int]]:
    """Return inclusive Excel row ranges for each contiguous display-key group."""
    if not rows:
        return []

    groups: List[Tuple[int, int]] = []
    group_start = 2
    current_key = rows[0][0]
    for index in range(1, len(rows)):
        if rows[index][0] != current_key:
            groups.append((group_start, index + 1))
            group_start = index + 2
            current_key = rows[index][0]
    groups.append((group_start, len(rows) + 1))
    return groups


def _merge_display_key_column(worksheet: Worksheet, rows: Sequence[CellRow]) -> None:
    """Merge contiguous identical display-key cells in column A."""
    for start_row, end_row in _iter_display_key_groups(rows):
        if end_row > start_row:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=end_row,
                end_column=1,
            )


def _apply_row_border(
    worksheet: Worksheet,
    row_index: int,
    border: Border,
) -> None:
    """Apply a border style across columns A-C on one row."""
    for column in range(1, 4):
        worksheet.cell(row=row_index, column=column).border = border  # noqa


def _apply_section_borders(worksheet: Worksheet, rows: Sequence[CellRow]) -> None:
    """Add top/bottom borders for display-key groups; header keeps top only.

    Header row uses freeze panes instead of a bottom border. Group borders
    span columns A-C so each key block reads as a full-width section.
    """
    _apply_row_border(worksheet, 1, TOP_BORDER)

    for start_row, end_row in _iter_display_key_groups(rows):
        if start_row == end_row:
            _apply_row_border(worksheet, start_row, TOP_BOTTOM_BORDER)
            continue
        _apply_row_border(worksheet, start_row, TOP_BORDER)
        _apply_row_border(worksheet, end_row, BOTTOM_BORDER)


def _format_worksheet(
    worksheet: Worksheet,
    rows: Sequence[CellRow],
) -> None:
    """Apply fonts, fills, widths, key alignment, freeze, and borders."""
    row_count = len(rows)
    for column, width in COL_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"  # noqa

    for column in range(1, 4):
        worksheet.cell(row=1, column=column).font = BOLD_FONT  # noqa

    for row_index in range(2, row_count + 2):
        key_cell = worksheet.cell(row=row_index, column=1)
        key_cell.font = BOLD_FONT  # noqa
        key_cell.alignment = KEY_ALIGNMENT  # noqa

        np_cell = worksheet.cell(row=row_index, column=2)
        c1_cell = worksheet.cell(row=row_index, column=3)
        np_cell.font = NORMAL_FONT  # noqa
        c1_cell.font = NORMAL_FONT  # noqa

        np_val = np_cell.value
        c1_val = c1_cell.value
        if np_val and c1_val and np_val == c1_val:
            np_cell.fill = MATCH_FILL  # noqa
            c1_cell.fill = MATCH_FILL  # noqa

    _apply_section_borders(worksheet, rows)


def write_excel(output_path: str, rows: Sequence[CellRow]) -> None:
    """Write comparison rows to an Excel workbook."""
    if os.path.exists(output_path):
        os.remove(output_path)

    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    worksheet = workbook.create_sheet(SHEET_TITLE)

    worksheet["A1"] = ""
    worksheet["B1"] = "NP1PP"
    worksheet["C1"] = "C1Y"

    for row_index, (key, np_val, c1_val) in enumerate(rows, start=2):
        worksheet.cell(row=row_index, column=1, value=key)
        if np_val:
            worksheet.cell(row=row_index, column=2, value=np_val)
        if c1_val:
            worksheet.cell(row=row_index, column=3, value=c1_val)

    _merge_display_key_column(worksheet, rows)
    _format_worksheet(worksheet, rows)
    workbook.save(output_path)


# #### Main
def main(argv: Optional[Sequence[str]] = None) -> int:  # pylint: disable=too-many-locals
    """Run the cell comparison pipeline."""
    args = parse_args(argv)

    np1_cells = read_list(args.np_file)
    c1y_cells = read_list(args.c1_file)
    np_set = set(np1_cells)
    c1_set = set(c1y_cells)
    all_cells = sorted(np_set | c1_set)
    print(
        f"NP1PP: {len(np1_cells)}, C1Y: {len(c1y_cells)}, "
        f"Combined unique: {len(all_cells)}"
    )

    if args.dump_suggested_keys:
        keys, table_path = dump_suggested_keys(
            all_cells, args.dump_suggested_keys
        )
        print(
            f"Wrote {len(keys)} KEY\\t中文 rows to {table_path} "
            "(one file for match order + Chinese)"
        )
        return 0

    function_keys, zh_map, table_note = resolve_function_key_config(
        args.function_keys,
        args.function_keys_file,
        args.function_key_zh_file,
    )
    patterns = split_format_filter(args.format_filter)
    replace_mappings = parse_format_replace(args.format_replace)
    regex_filters = compile_regex_filters(patterns)
    regex_replaces = compile_regex_replaces(replace_mappings)
    print(f"Function keys: {len(function_keys)}")
    if table_note:
        print(f"Key/Chinese table: {table_note}")
    print(
        f"Chinese map entries: "
        f"{len(zh_map) if zh_map else '(builtin fallback)'}"
    )
    print(f"Regex filters (ordered): {patterns or '(none)'}")
    print(f"Regex replaces (ordered): {replace_mappings or '(none)'}")

    groups, unmatched = group_cells(
        all_cells,
        np_set,
        c1_set,
        function_keys,
        regex_filters,
        regex_replaces,
    )
    rows = build_rows(groups, zh_map or None)
    print(
        f"Output rows: {len(rows)}, Unique function keys used: {len(groups)}, "
        f"Unmatched cells: {len(unmatched)}"
    )
    if unmatched:
        preview = ", ".join(unmatched[:10])
        more = "" if len(unmatched) <= 10 else f", ... (+{len(unmatched) - 10})"
        print(
            f"Warning: unmatched cells keep post-filter names as keys: "
            f"{preview}{more}",
            file=sys.stderr,
        )

    write_excel(args.output, rows)
    print(f"Done! Saved to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
